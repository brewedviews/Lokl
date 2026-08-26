"""xlsx template + parser for merchant bulk product upload.

The template uses Excel data-validation dropdowns so merchants can't fat-finger an L1
or L2 name — every cell is constrained to a known value. The Returnable column is also
a Yes/No dropdown. Sizes + stock are typed manually because they are per-product.
"""
from __future__ import annotations

import io
import re
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from seed_data import L1_CATEGORIES, L2_BY_L1

L1_NAMES = [c["name"] for c in L1_CATEGORIES]
L1_NAME_TO_ID = {c["name"].lower(): c["id"] for c in L1_CATEGORIES}
L2_NAME_TO_ID: dict[tuple[str, str], str] = {}
for lid, subs in L2_BY_L1.items():
    for s in subs:
        L2_NAME_TO_ID[(lid, s["name"].lower())] = s["id"]
ALL_L2_NAMES = sorted({s["name"] for subs in L2_BY_L1.values() for s in subs})
GENDERS = ["women", "men", "unisex", "kids"]
YES_NO = ["Yes", "No"]

HEADERS = [
    "product name", "product description",
    "l1 category", "l2 category", "gender",
    "mrp", "selling price",
    "sizes", "stock_per_size",
    "returnable", "return window (hours)", "try & buy",
    "brand",
]
# 10 example rows the merchant can replace.
EXAMPLES = [
    ["Indigo Block-Print Kurta", "Pure cotton hand-block",
     "Women", "Kurtas & Suits", "", 3499, 1899, "S;M;L;XL", "50;100;39;10", "Yes", 24, "No"],
    ["Oversized Tee", "240GSM oversized graphic tee",
     "Men", "T-Shirts", "", 1499, 899, "M;L;XL", "30;45;20", "Yes", 24, "No"],
    ["White Court Sneakers", "Classic low-top court sneakers",
     "Footwear", "Casual Shoes", "", 4999, 3499, "7;8;9;10", "8;12;10;6", "No", "", "Yes"],
]


def _add_dropdown(ws, col_letter: str, options: list[str], first_row: int = 2, last_row: int = 200):
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")
    ws.add_data_validation(dv)


def _l1_range_name(l1_name: str) -> str:
    """Excel defined names can't contain spaces/most punctuation and can't
    start with a digit — sanitize the L1 display name into a stable,
    collision-free range name (e.g. "Lingerie & Innerwear" -> L2_Lingerie___Innerwear)."""
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", l1_name)
    if cleaned[:1].isdigit():
        cleaned = "_" + cleaned
    return f"L2_{cleaned}"


def _add_l2_dependent_dropdown(wb, ws, l1_col: str, l2_col: str, first_row: int = 2, last_row: int = 200):
    """Real Excel-native L1→L2 dependent dropdown (G12 P1-9) — replaces the
    old flat, unfiltered "every L2 in the whole taxonomy" list, which let a
    merchant pick e.g. "Sarees" under L1 "Men" with zero warning until the
    server silently skipped the row. Standard technique: one named range per
    L1 on a hidden helper sheet, L2 column validated via an INDIRECT formula
    keyed off that row's own L1 cell — so the visible dropdown itself only
    ever offers L2 values valid for whatever L1 the merchant already chose.
    """
    helper = wb.create_sheet("L2Lists")
    helper.sheet_state = "hidden"
    for col_idx, l1 in enumerate(L1_CATEGORIES, start=1):
        subs = L2_BY_L1.get(l1["id"], [])
        col_letter = get_column_letter(col_idx)
        helper.cell(row=1, column=col_idx, value=l1["name"])
        if not subs:
            # No L2s for this L1 — still define a (single blank-cell) named
            # range so INDIRECT resolves to something valid rather than a
            # #REF! error when the merchant opens the L2 dropdown for it.
            helper.cell(row=2, column=col_idx, value=None)
            ref = f"'L2Lists'!${col_letter}$2:${col_letter}$2"
        else:
            for i, s in enumerate(subs, start=2):
                helper.cell(row=i, column=col_idx, value=s["name"])
            ref = f"'L2Lists'!${col_letter}$2:${col_letter}${len(subs) + 1}"
        wb.defined_names[_l1_range_name(l1["name"])] = DefinedName(_l1_range_name(l1["name"]), attr_text=ref)

    # Excel formulas can't regex-replace arbitrary characters the way
    # `_l1_range_name` does in Python — this chains one SUBSTITUTE per
    # special character actually present in today's L1 names (space, &).
    # A future L1 name introducing a new punctuation character needs a
    # matching SUBSTITUTE added here to stay in sync with `_l1_range_name`.
    formula = f'INDIRECT("L2_"&SUBSTITUTE(SUBSTITUTE(${l1_col}{first_row}," ","_"),"&","_"))'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False, showErrorMessage=False)
    dv.add(f"{l2_col}{first_row}:{l2_col}{last_row}")
    ws.add_data_validation(dv)


def build_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    header_fill = PatternFill("solid", fgColor="1A2B4C")
    header_font = Font(bold=True, color="FFFFFF")
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 26
    widths = [28, 30, 14, 22, 12, 10, 10, 22, 22, 12, 18, 12, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r, row in enumerate(EXAMPLES, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    # Dropdowns
    _add_dropdown(ws, "C", L1_NAMES)
    _add_l2_dependent_dropdown(wb, ws, l1_col="C", l2_col="D")
    _add_dropdown(ws, "E", GENDERS)
    _add_dropdown(ws, "J", YES_NO)
    _add_dropdown(ws, "L", YES_NO)

    # Instructions sheet
    info = wb.create_sheet("How to fill")
    info["A1"] = "Lokl bulk product upload — instructions"
    info["A1"].font = Font(bold=True, size=14)
    info_lines = [
        "",
        "• Fill one row per product on the 'Products' sheet.",
        "• Click into the l1 category cell first — a DROPDOWN arrow appears on the right.",
        "  Then click into l2 category: its dropdown automatically shows ONLY the sub-categories",
        "  valid for whatever l1 category you picked on that row. Pick l1 before l2, or the l2",
        "  dropdown will still be showing the previous row's options.",
        "  You cannot type custom categories: only listed values are accepted.",
        "• See the 'L1 → L2 reference' sheet for the full category tree at a glance.",
        "• sizes: comma or semicolon-separated, e.g.  S,M,L,XL  or  S;M;L;XL  or  7;8;9;10",
        "• stock_per_size: same count as sizes, e.g. 50,100,39,10 — quantity per size.",
        "• returnable: Yes / No. Defaults to No if blank. (Innerwear, perishables: keep No.)",
        "• return window (hours): only used when returnable is Yes. Whole number, 1-24. Defaults to 24 if blank.",
        "• try & buy: Yes / No — can the customer try this on at the door and only pay for what they keep?",
        "  Defaults to No if blank.",
        "• brand: optional. Matches an existing brand by name (case-insensitive) from Lokl's brand list.",
        "  Brands can't be created from this sheet — an unrecognized name is noted in the upload summary",
        "  and the product is still created, just without a brand tag. Check spelling or ask Lokl to add it.",
        "• mrp / price are in INR and must be positive numbers.",
        "• After upload, every row appears in Products as a draft — add images & tweak before go-live.",
    ]
    for i, line in enumerate(info_lines, start=2):
        info.cell(row=i, column=1, value=line)
    info.column_dimensions["A"].width = 100

    # L1 → L2 reference table — gives merchants the full category tree at a glance
    ref = wb.create_sheet("L1 → L2 reference")
    ref["A1"] = "Category (L1)"
    ref["B1"] = "Sub-category (L2)"
    for c in ("A1", "B1"):
        ref[c].font = header_font
        ref[c].fill = header_fill
        ref[c].alignment = Alignment(vertical="center")
    ref.row_dimensions[1].height = 24
    ref.column_dimensions["A"].width = 24
    ref.column_dimensions["B"].width = 38
    row = 2
    for c in L1_CATEGORIES:
        subs = L2_BY_L1.get(c["id"], [])
        if not subs:
            ref.cell(row=row, column=1, value=c["name"])
            ref.cell(row=row, column=2, value="(no sub-category — leave l2 category blank)").font = Font(italic=True, color="888888")
            row += 1
        else:
            for s in subs:
                ref.cell(row=row, column=1, value=c["name"])
                ref.cell(row=row, column=2, value=s["name"])
                row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()



# ============================================================================
# G14 — universal column mapping. ONE canonical alias table, used by BOTH the
# xlsx path (parse_uploaded_xlsx below) and the csv path (server.py's
# bulk_products) — previously csv never got _HEADER_ALIAS treatment at all
# and relied solely on _row_to_product's own narrower inline `or`-chains, a
# real "separate business logic for CSV" architecture problem. Every alias
# list below matches the "SUPPORTED COLUMN MAPPING" list from the G14 brief
# verbatim — deliberately NOT a fuzzy/heuristic matcher, so a column that
# isn't on this list is reported as unmapped rather than guessed at.
#
# "image" is handled specially below (map_row_headers) since a file can
# legitimately have several image columns (Image 1/2/3...) — those all need
# distinct canonical keys, not one that collides.
# ============================================================================
CANONICAL_ALIASES: dict[str, list[str]] = {
    "name": ["product name", "product_name", "product", "name", "item name", "item"],
    "description": ["description", "product description", "product_description", "details"],
    "gender": ["gender", "sex", "target gender"],
    "l1": ["l1", "l1 category", "l1_category", "category", "main category", "department"],
    "l2": ["l2", "l2 category", "l2_category", "sub category", "subcategory", "sub-category", "l2 sub-category"],
    "mrp": ["mrp", "mrp price", "maximum retail price", "list price"],
    "price": ["selling price", "selling_price", "sale price", "selling price (inr)", "price", "offer price"],
    "stock_total": ["stock", "quantity", "inventory", "available stock"],
    "sizes": ["sizes", "size", "available sizes"],
    "stock_per_size": ["stock per size", "stock_per_size", "stock by size", "size stock", "quantity per size"],
    "returnable": ["returnable", "return eligible", "return allowed", "is returnable"],
    "return_window_hours": ["return window", "return window hours", "return window (hours)", "return_window_hours", "return hours"],
    "try_at_doorstep": ["try & buy", "try and buy", "try & buy available", "try at doorstep", "try at doorstep available", "try_at_doorstep"],
    "brand": ["brand", "brand name", "brand_name"],
}
_IMAGE_ALIASES = {"image", "image url", "product image", "product image url"}

# Flat lowercased-alias -> canonical-field lookup, built once from the table
# above. Longer/more-specific aliases never lose to shorter ones since every
# entry maps to exactly one field — no ambiguity to resolve.
_FLAT_ALIAS: dict[str, str] = {
    alias: field for field, aliases in CANONICAL_ALIASES.items() for alias in aliases
}

# Fields a sellable product genuinely can't exist without — drives both the
# detect endpoint's `unmapped_required` and (indirectly, via _row_to_product
# already requiring them) row-level validation. `l2`/`gender` are NOT always
# required (gender-neutral L1s per G12) so they're intentionally excluded —
# the row-level validator still enforces l2-when-the-L1-has-l2s.
REQUIRED_CANONICAL_FIELDS = ["name", "l1", "price", "mrp"]


def _is_image_header(h: str) -> bool:
    if h in _IMAGE_ALIASES:
        return True
    # "image 1", "image 2", "image_3"... — numbered variants of the same alias.
    base = re.sub(r"[\s_]*\d+$", "", h).strip()
    return base in _IMAGE_ALIASES


def map_row_headers(
    raw_headers: list[str], overrides: dict[str, str | None] | None = None,
) -> tuple[list[str | None], list[dict]]:
    """Map a file's raw header row to Lokl's canonical product fields.

    `overrides` (G14 §4) — an optional {lowercased header: canonical_field}
    map from the frontend's mapping/preview step, for the rare column the
    automatic aliasing below can't confidently place. Checked FIRST for
    each header, so a merchant's manual choice always wins over (or can
    explicitly un-map, via a null value) the automatic guess.

    Returns (canonical_keys, columns_report):
      - canonical_keys[i] is the canonical field name for column i (or None
        if unmapped) — image columns get "image", "image_2", "image_3"...
        in encounter order so multiple image cells never collide.
      - columns_report is [{"header": <original text>, "mapped_field": ...}]
        for every column, in order — this is exactly what the /bulk/detect
        endpoint hands the frontend for the mapping/preview step.

    Confidently-mapped only: a header not on CANONICAL_ALIASES / the image
    alias set maps to None (shown to the merchant as "Not mapped") rather
    than guessed at — see this module's own top-of-section comment.
    """
    canonical_keys: list[str | None] = []
    columns_report: list[dict] = []
    image_slot = 0
    for raw in raw_headers:
        h = (raw or "").strip().lower()
        if not h:
            canonical_keys.append(None)
            columns_report.append({"header": raw or "", "mapped_field": None})
            continue
        if overrides is not None and h in overrides:
            field = overrides[h]
            canonical_keys.append(field)
            columns_report.append({"header": raw, "mapped_field": field})
            continue
        if _is_image_header(h):
            image_slot += 1
            key = "image" if image_slot == 1 else f"image_{image_slot}"
            canonical_keys.append(key)
            columns_report.append({"header": raw, "mapped_field": "image"})
            continue
        field = _FLAT_ALIAS.get(h)
        canonical_keys.append(field)
        columns_report.append({"header": raw, "mapped_field": field})
    return canonical_keys, columns_report


def looks_like_lokl_template(raw_headers: list[str]) -> bool:
    """True only when the header row is exactly today's generated template's
    header text (case-insensitive) — the ONE signal allowed to skip the
    frontend's mapping/preview step per the brief's explicit "do not force a
    mapping screen when the file already matches canonical structure." Never
    used to gate validation itself — only the UI's mapping-screen decision."""
    cleaned = [(h or "").strip().lower() for h in raw_headers if (h or "").strip()]
    return cleaned == HEADERS


def _cell_to_str(v) -> str:
    """Stringify a raw spreadsheet cell without mangling Excel's numeric
    typing. Excel stores a merchant's plain "6" as the float 6.0 the moment
    it looks numeric — `str(6.0)` is "6.0", which then fails to match a size
    like "6" downstream. A whole-number float/int becomes its plain integer
    string; anything else (already a string, e.g. "6;7") is left as-is."""
    if v is None:
        return ""
    if isinstance(v, bool):
        return str(v)
    if isinstance(v, int):
        return str(v)
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    return str(v).strip()


def parse_uploaded_xlsx(
    data: bytes, sheet: str | None = None, overrides: dict[str, str | None] | None = None,
) -> tuple[list[dict], dict]:
    """Parse a merchant-uploaded xlsx into canonical row dicts.

    Sheet selection (G14 §12): an explicit `sheet` name wins if given and
    exists; else our own template's "Products" sheet name wins if present
    (back-compat); else, among VISIBLE sheets only (hidden helper sheets
    like the template's own "L2Lists" are never candidates), the sheet whose
    header row has the most columns this module can confidently map is
    chosen; else the workbook's own active sheet.

    Returns (rows, meta) — meta is the same "detect" info the /bulk/detect
    endpoint surfaces: {sheet_names, selected_sheet, columns, row_count,
    looks_like_lokl_template}. Row dicts are keyed by CANONICAL field names
    (already alias-mapped via map_row_headers) plus `_row_num` — the real,
    1-indexed spreadsheet row a merchant would see in Excel.
    """
    wb = load_workbook(io.BytesIO(data), data_only=True)
    visible_sheets = [s for s in wb.sheetnames if wb[s].sheet_state == "visible"]

    def _header_row(name: str) -> list[str]:
        ws = wb[name]
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        return [str(h).strip() if h is not None else "" for h in first]

    if sheet and sheet in wb.sheetnames:
        selected = sheet
    elif "Products" in wb.sheetnames:
        selected = "Products"
    elif visible_sheets:
        scored = [(s, sum(1 for k in map_row_headers(_header_row(s))[0] if k)) for s in visible_sheets]
        scored.sort(key=lambda t: t[1], reverse=True)
        selected = scored[0][0] if scored and scored[0][1] > 0 else (wb.active.title if wb.active else visible_sheets[0])
    else:
        selected = wb.active.title if wb.active else wb.sheetnames[0]

    ws = wb[selected]
    raw_headers = _header_row(selected)
    canonical_keys, columns_report = map_row_headers(raw_headers, overrides)

    rows: list[dict] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue
        if not any(v not in (None, "") for v in row):
            continue
        d: dict = {}
        for i, key in enumerate(canonical_keys):
            if not key or i >= len(row):
                continue
            d[key] = row[i]
        d["_row_num"] = row_idx
        rows.append(d)

    meta = {
        "sheet_names": visible_sheets or wb.sheetnames,
        "selected_sheet": selected,
        "columns": columns_report,
        "row_count": len(rows),
        # Header-content match only (not gated on sheet name) — a merchant
        # file that happens to already use Lokl's exact canonical headers
        # is just as unambiguous as our own generated template, and the
        # brief is explicit: don't force a mapping screen on an
        # already-canonical file, regardless of whose file it is.
        "looks_like_lokl_template": looks_like_lokl_template(raw_headers),
    }
    return rows, meta
