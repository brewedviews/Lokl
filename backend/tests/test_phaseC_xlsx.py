"""Phase C — xlsx template + bulk upload."""
import os, uuid, io, pytest, requests
from pathlib import Path
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv(Path(__file__).resolve().parents[2] / "frontend" / ".env")
BASE = os.environ["REACT_APP_BACKEND_URL"].rstrip("/")
API = BASE + "/api"


def _admin_token():
    r = requests.post(f"{API}/admin/login", json={"email": "admin@lokl.in", "password": "Admin@2026"}, timeout=30)
    return r.json()["token"]


@pytest.fixture(scope="module")
def merchant_token():
    email = f"phaseC_{uuid.uuid4().hex[:8]}@lokl.in"
    phone = f"+9199{str(uuid.uuid4().int)[:8]}"
    r = requests.post(f"{API}/auth/register", json={
        "email": email, "password": "PhaseC@2026",
        "store_name": f"PhaseC {uuid.uuid4().hex[:4]}",
        "owner_name": "Owner", "phone": phone, "city": "Bhilai",
    }, timeout=30)
    tok = r.json()["token"]
    mid = r.json()["merchant"]["id"]
    # KYC + approve so bulk endpoint passes the KYC gate
    requests.post(f"{API}/merchant/kyc/submit", headers={"Authorization": f"Bearer {tok}"}, json={
        "pan_number": "ABCDE1234F", "business_name": "B", "business_category": "Multi-category",
        "business_type": "Proprietorship", "business_address": "Bhilai",
        "bank_account_number": "1", "bank_ifsc": "S", "account_holder_name": "X",
        "pan_doc_b64": "x", "cancelled_cheque_b64": "y",
    }, timeout=30)
    atok = _admin_token()
    requests.post(f"{API}/admin/merchants/{mid}/approve",
                  headers={"Authorization": f"Bearer {atok}"}, timeout=30)
    requests.post(f"{API}/merchant/storefront", headers={"Authorization": f"Bearer {tok}"}, json={
        "tagline": "x", "story": "x", "specialties": [], "banner": "x", "banners": ["x"],
        "address": "Bhilai", "area": "Sec", "locality": "Sec", "city": "Bhilai",
        "opens_at": "10:00", "closes_at": "18:00", "timing": "10:00 - 18:00",
    }, timeout=30)
    return tok


def test_template_xlsx_has_dropdowns_and_returnable(merchant_token):
    r = requests.get(f"{API}/merchant/products/template.xlsx",
                     headers={"Authorization": f"Bearer {merchant_token}"}, timeout=30)
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers.get("content-type", "")
    wb = load_workbook(io.BytesIO(r.content))
    assert "Products" in wb.sheetnames
    ws = wb["Products"]
    headers = [ws.cell(row=1, column=i).value for i in range(1, 11)]
    assert "returnable" in headers, f"returnable column missing: {headers}"
    assert "l1" in headers and "l2" in headers
    dvs = list(ws.data_validations.dataValidation)
    # Expect 4 dropdowns: L1, L2, gender, returnable
    assert len(dvs) == 4, f"expected 4 data validations got {len(dvs)}"


def test_bulk_xlsx_round_trip(merchant_token):
    r = requests.get(f"{API}/merchant/products/template.xlsx",
                     headers={"Authorization": f"Bearer {merchant_token}"}, timeout=30)
    files = {"file": ("lokl-products-template.xlsx", r.content,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r2 = requests.post(f"{API}/merchant/products/bulk", files=files,
                       headers={"Authorization": f"Bearer {merchant_token}"}, timeout=60)
    assert r2.status_code == 200, r2.text
    body = r2.json()
    assert body["created"] == 3
    assert isinstance(body.get("created_ids"), list) and len(body["created_ids"]) == 3
    # All bulk-uploaded products should be paused (need image) until merchant adds image / toggles go-live
    items = requests.get(f"{API}/merchant/products",
                        headers={"Authorization": f"Bearer {merchant_token}"}, timeout=30).json()
    for pid in body["created_ids"]:
        p = next(x for x in items if x["id"] == pid)
        assert p.get("paused") is True, f"{pid} should be paused on bulk upload"
        assert p.get("needs_image") is True
    # `return_eligible` should reflect the Yes/No column
    by_name = {p["name"]: p for p in items}
    assert by_name["Indigo Block-Print Kurta"]["return_eligible"] is True
    assert by_name["White Court Sneakers"]["return_eligible"] is False


def test_bulk_legacy_csv_still_works(merchant_token):
    csv_text = b"""name,description,l1,l2,gender,mrp,price,sizes,stock_per_size,returnable
Test CSV Tee,short desc,Men,T-shirts,,999,599,M;L,10;5,Yes
"""
    r = requests.post(f"{API}/merchant/products/bulk",
                      files={"file": ("x.csv", csv_text, "text/csv")},
                      headers={"Authorization": f"Bearer {merchant_token}"}, timeout=30)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["created"] >= 1
